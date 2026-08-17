from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from employment_ai.core.context import AppContext
from employment_ai.core.contracts import Plugin, PluginManifest


def _iso(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value or "").strip()


def _split(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    return [item.strip() for item in str(value).split("|") if item.strip()]


def _integer(value: Any) -> int:
    if isinstance(value, bool):
        raise ValueError("boolean is not a valid integer")
    return int(float(value))


def _number(value: Any) -> float:
    return float(value)


class ExcelSourceService:
    def __init__(self, context: AppContext) -> None:
        self.context = context

    @staticmethod
    def _validate_container(path: Path) -> None:
        try:
            with ZipFile(path) as archive:
                entries = archive.infolist()
                if len(entries) > 250:
                    raise ValueError("xlsx压缩包文件项过多")
                total_size = sum(item.file_size for item in entries)
                if total_size > 80 * 1024 * 1024:
                    raise ValueError("xlsx解压后体积超过80MB安全限制")
                for item in entries:
                    if item.flag_bits & 0x1:
                        raise ValueError("不接受加密xlsx")
                    if item.file_size > 0 and item.compress_size > 0:
                        if item.file_size / item.compress_size > 250:
                            raise ValueError("xlsx压缩比异常")
                    parts = Path(item.filename).parts
                    if item.filename.startswith("/") or ".." in parts:
                        raise ValueError("xlsx包含不安全路径")
        except BadZipFile as exc:
            raise ValueError("文件不是有效的xlsx压缩容器") from exc

    @staticmethod
    def _rows(sheet: Any) -> tuple[list[str], list[dict[str, Any]], list[dict[str, Any]]]:
        iterator = sheet.iter_rows()
        header_cells = next(iterator, ())
        headers = [str(cell.value or "").strip() for cell in header_cells]
        formula_errors: list[dict[str, Any]] = []
        rows: list[dict[str, Any]] = []
        for row_no, cells in enumerate(iterator, start=2):
            if all(cell.value in (None, "") for cell in cells):
                continue
            record: dict[str, Any] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                cell = cells[index] if index < len(cells) else None
                if cell is not None and cell.data_type == "f":
                    formula_errors.append(
                        {
                            "sheet": sheet.title,
                            "row": row_no,
                            "field": header,
                            "code": "FORMULA_NOT_ALLOWED",
                            "message": "导入快照不允许公式单元格",
                            "severity": "error",
                        }
                    )
                record[header] = cell.value if cell is not None else None
            rows.append(record)
        return headers, rows, formula_errors

    @staticmethod
    def _normalize_person(row: dict[str, Any]) -> dict[str, Any]:
        return {
            "person_id": str(row.get("person_id", "")).strip(),
            "education": str(row.get("education", "")).strip(),
            "major": str(row.get("major", "") or "").strip(),
            "skill_level": str(row.get("skill_level", "无") or "无").strip(),
            "skills": _split(row.get("skills")),
            "employment_status": str(row.get("employment_status", "")).strip(),
            "expected_salary_min": _integer(row.get("expected_salary_min", 0)),
            "expected_salary_max": _integer(row.get("expected_salary_max", 0)),
            "preferred_region": str(row.get("preferred_region", "")).strip(),
            "preferred_industries": _split(row.get("preferred_industries")),
            "special_tags": _split(row.get("special_tags")),
            "town": str(row.get("town", "")).strip(),
            "village": str(row.get("village", "") or "").strip(),
            "years_experience": _number(row.get("years_experience", 0)),
            "available_shift": str(row.get("available_shift", "")).strip(),
            "source_updated_at": _iso(row.get("source_updated_at")),
        }

    @staticmethod
    def _normalize_job(row: dict[str, Any]) -> dict[str, Any]:
        return {
            "job_id": str(row.get("job_id", "")).strip(),
            "employer_name": str(row.get("employer_name", "")).strip(),
            "job_title": str(row.get("job_title", "")).strip(),
            "region": str(row.get("region", "")).strip(),
            "salary_min": _integer(row.get("salary_min", 0)),
            "salary_max": _integer(row.get("salary_max", 0)),
            "education_min": str(row.get("education_min", "")).strip(),
            "experience_min": _number(row.get("experience_min", 0)),
            "required_skills": _split(row.get("required_skills")),
            "industry": str(row.get("industry", "")).strip(),
            "shift": str(row.get("shift", "")).strip(),
            "headcount": _integer(row.get("headcount", 0)),
            "valid_until": _iso(row.get("valid_until"))[:10],
            "status": str(row.get("status", "")).strip(),
        }

    def import_workbook(self, path: Path, filename: str, actor: str) -> dict[str, Any]:
        self._validate_container(path)
        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
        except Exception as exc:
            raise ValueError(f"无法读取xlsx文件：{exc}") from exc

        required_sheets = {"person_snapshot", "job_snapshot"}
        missing = required_sheets.difference(workbook.sheetnames)
        if missing:
            errors = [
                {
                    "sheet": name,
                    "row": 1,
                    "field": "*",
                    "code": "MISSING_SHEET",
                    "message": f"缺少工作表：{name}",
                    "severity": "error",
                }
                for name in sorted(missing)
            ]
            batch_id = self.context.db.create_batch(filename, "rejected", 0, 0, errors)
            self.context.audit("batch.rejected", actor, {"batch_id": batch_id, "errors": errors})
            return {"batch_id": batch_id, "status": "rejected", "errors": errors}

        _, people_raw, people_formula_errors = self._rows(workbook["person_snapshot"])
        _, jobs_raw, jobs_formula_errors = self._rows(workbook["job_snapshot"])
        try:
            people = [self._normalize_person(row) for row in people_raw]
            jobs = [self._normalize_job(row) for row in jobs_raw]
        except (TypeError, ValueError) as exc:
            errors = [
                {
                    "sheet": "workbook",
                    "row": 0,
                    "field": "*",
                    "code": "TYPE_CONVERSION",
                    "message": f"字段类型转换失败：{exc}",
                    "severity": "error",
                }
            ]
            batch_id = self.context.db.create_batch(filename, "rejected", 0, 0, errors)
            return {"batch_id": batch_id, "status": "rejected", "errors": errors}

        quality = self.context.services.get("data.quality").validate(people, jobs)
        self.context.events.publish(
            "data.quality.checked",
            {
                "filename": filename,
                "people": len(people),
                "jobs": len(jobs),
                "errors": len(quality.errors),
                "warnings": len(quality.warnings),
            },
        )
        errors = people_formula_errors + jobs_formula_errors + quality.errors
        warnings = quality.warnings
        status = "ready" if not errors else "rejected"
        batch_id = self.context.db.create_batch(
            filename, status, len(people), len(jobs), errors + warnings
        )
        if not errors:
            normalizer = self.context.services.get("semantic.normalize")
            people = [normalizer.normalize_person(person) for person in people]
            jobs = [normalizer.normalize_job(job) for job in jobs]
            self.context.db.store_snapshot(batch_id, people, jobs)
            self.context.events.publish(
                "semantic.normalized",
                {
                    "batch_id": batch_id,
                    "people": len(people),
                    "jobs": len(jobs),
                    "ontology_version": normalizer.version,
                },
            )
            self.context.events.publish(
                "snapshot.imported",
                {"batch_id": batch_id, "people": len(people), "jobs": len(jobs)},
            )
            self.context.audit(
                "batch.imported",
                actor,
                {
                    "batch_id": batch_id,
                    "filename": filename,
                    "people": len(people),
                    "jobs": len(jobs),
                    "warnings": len(warnings),
                },
            )
        else:
            self.context.audit(
                "batch.rejected",
                actor,
                {"batch_id": batch_id, "filename": filename, "error_count": len(errors)},
            )
        return {
            "batch_id": batch_id,
            "status": status,
            "person_count": len(people),
            "job_count": len(jobs),
            "errors": errors,
            "warnings": warnings,
        }


class ExcelSourcePlugin(Plugin):
    manifest = PluginManifest(
        id="excel-source",
        version="1.0.0",
        name="Excel数据源",
        description="读取双Sheet受控快照，通过质量门后形成不可变批次",
        provides=("data.source.excel",),
        requires=("data.quality", "semantic.normalize"),
        permissions=("file:read:xlsx", "snapshot:write"),
        events_out=("snapshot.imported",),
        cleanup_strategy="remove importer service; preserve batch evidence and audit records",
    )

    def install(self, context: AppContext) -> None:
        context.services.register(
            "data.source.excel", self.manifest.id, ExcelSourceService(context)
        )
